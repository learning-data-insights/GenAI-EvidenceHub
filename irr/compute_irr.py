#!/usr/bin/env python3
"""
compute_irr.py — Inter-rater reliability (Krippendorff's alpha) for the
GenAI Evidence Hub coding workbook.

Reverse-engineered to match the team's "human-calculated" numbers:

  * metric          : Krippendorff's alpha, nominal level of measurement
  * reliability unit: one (Paper ID, Research Question ID, Field) cell
  * observers       : the coders in the `Reviewer` column
  * computation     : ALL cells pooled into ONE coincidence matrix -> ONE alpha
                      (overall = all cells; per-paper = cells of that paper;
                       pairwise = restrict observers to two coders)
  * fields          : taken from the workbook's `FieldList+IRR` sheet
                      (those marked "Y"), EXCLUDING free-text prose fields
                      that can't be auto-scored (see EXCLUDE_FREETEXT) and
                      with numeric fields normalized so 85 == 85.0
  * missing         : blank / "N/A" cells are dropped; a unit needs >=2 coders

Usage:
  python compute_irr.py WORKBOOK.xlsx --sheet SHEET_NAME
  python compute_irr.py WORKBOOK.xlsx --sheet SHEET_NAME --per-paper --pairs --csv out.csv
  python compute_irr.py WORKBOOK.xlsx --all-sheets
  python compute_irr.py WORKBOOK.xlsx --all-sheets --aliases aliases.json

Can also be imported as a library; see load_workbook_records() and
krippendorff_alpha().
"""
import argparse, csv, math, re, sys
from collections import defaultdict, Counter
from itertools import combinations

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

# ---------------------------------------------------------------- config ----
# Free-text fields the team scores by "concept agreement", NOT exact match.
# Excluded from automated alpha (matched by normalized name).
EXCLUDE_FREETEXT = {
    "research question",
    "tested llm model & version used",
    "tested model & techniques list",
    "baseline list",
}
# Reviewer-name aliases (lowercased spelling -> canonical name). This default
# covers one team's coders; any name NOT listed here passes through unchanged,
# so the tool already works for arbitrary coders. Supply --aliases PATH (a JSON
# object of {"spelling": "Canonical"}) to add/override entries without editing
# this file -- needed only when one person appears under multiple spellings
# (e.g. initials vs. full name).
NAME_MAP = {
    "jw": "John", "john": "John", "john w": "John",
    "mb": "Maggie", "maggie": "Maggie",
    "alexander white": "Alexander", "alexanderwhite": "Alexander",
    "alexander": "Alexander", "chris": "Chris", "aaron": "Aaron",
    "heeryung": "Heeryung", "warren": "Warren", "alexis": "Alexis",
    "nidhi": "Nidhi", "steadman": "Steadman",
}
NON_REVIEWER = {"reviewer", "string", "selection", "multiselection",
                "n/a", "true", "false", "numerical"}
MISSING_TOKENS = {"", "n/a", "na", "none", "-", "--", "n.a.", "tbd", "?"}

# Fallback IRR field spec if the workbook has no FieldList+IRR sheet.
NONMETRIC_Y = {
    "domain", "ai innovation", "fairness acknowledged", "fairness evaluated",
    "research question", "total number of models tested",
    "tested llm model & version used", "tested model & techniques list",
    "baselines tested", "baseline list", "prompting techniques",
    "qualitative evaluations used",
}

# ------------------------------------------------------------ normalize ----
def norm_name(x, name_map=NAME_MAP):
    if x is None:
        return None
    s = str(x).strip()
    return name_map.get(s.lower(), s)

def load_aliases(path):
    """Load a JSON {"spelling": "Canonical"} map and merge it over the
    built-in NAME_MAP (caller-supplied entries win)."""
    import json
    with open(path) as fh:
        raw = json.load(fh)
    merged = dict(NAME_MAP)
    merged.update({str(k).strip().lower(): str(v).strip() for k, v in raw.items()})
    return merged

def norm_value(v, dtype):
    """Canonical token for a cell, or None if missing.
    Numeric fields are parsed/rounded so 85 == 85.0 == '85'."""
    if v is None:
        return None
    dt = (dtype or "").strip().lower()
    if isinstance(v, str) and v.strip().lower() in MISSING_TOKENS:
        return None
    if "numerical" in dt:
        num = _to_number(v)
        if num is not None:
            return _fmt_number(num)
        # non-numeric junk in a numeric field -> treat as missing
        s = str(v).strip().lower()
        return None if s in MISSING_TOKENS else s
    if isinstance(v, float):
        if math.isnan(v):
            return None
        return _fmt_number(v)
    s = str(v).strip()
    if s.lower() in MISSING_TOKENS:
        return None
    if "multiselection" in dt:
        # order-independent: sort the comma-separated selections
        parts = [p.strip().lower() for p in s.split(",") if p.strip()]
        return "|".join(sorted(parts)) if parts else None
    return s.lower()

def _to_number(v):
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    pct = s.endswith("%")
    s = s.rstrip("%").strip()
    m = re.fullmatch(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", s)
    if not m:
        return None
    val = float(s)
    return val / 100.0 if pct else val

def _fmt_number(x):
    x = round(float(x), 4)
    return str(int(x)) if x == int(x) else repr(x)

# --------------------------------------------------------------- loader ----
def _norm(s):
    return str(s).strip().lower() if s is not None else ""

def load_workbook_records(path, sheet, include_fields=None, name_map=NAME_MAP):
    """Return (records, fields_used) where records is a list of
    (paper, rq, field, coder, value). Missing values kept as None."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        sys.exit(f"sheet {sheet!r} not found. Available: {wb.sheetnames}")
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return [], []
    header = [(_norm(h)) for h in rows[0]]
    disp = [str(h).strip() if h is not None else "" for h in rows[0]]
    dtypes = [str(d).strip() if d is not None else "" for d in rows[1]]
    data = rows[2:]

    irr_y = include_fields if include_fields is not None else load_field_spec(wb)
    # column index of reviewer / paper / rq
    def first(name):
        for j, h in enumerate(header):
            if h == name:
                return j
        return None
    rcol = first("reviewer")
    pcol = first("paper id")
    qcol = first("research question id")
    if rcol is None or pcol is None:
        sys.exit(f"sheet {sheet!r}: missing Reviewer/Paper ID columns")

    # resolve the duplicate "Domain": keep the later (study-design) occurrence
    domain_idxs = [j for j, h in enumerate(header) if h == "domain"]
    domain_keep = max(domain_idxs) if domain_idxs else None

    field_cols = []  # (display_name, col_index, dtype)
    for j, h in enumerate(header):
        if not h:
            continue
        if h == "domain" and j != domain_keep:
            continue
        if not _is_included(h, irr_y):
            continue
        field_cols.append((disp[j], j, dtypes[j]))

    records = []
    for r in data:
        if not any(c is not None for c in r):
            continue
        coder = norm_name(r[rcol], name_map) if rcol < len(r) else None
        if coder is None or coder.lower() in NON_REVIEWER:
            continue
        if r[pcol] is None:
            continue
        paper = norm_value(r[pcol], "numerical")
        rq = norm_value(r[qcol], "numerical") if (qcol is not None and qcol < len(r) and r[qcol] is not None) else "_"
        for name, j, dt in field_cols:
            val = norm_value(r[j], dt) if j < len(r) else None
            records.append((paper, rq, name, coder, val))
    return records, [f[0] for f in field_cols]

def load_field_spec(wb):
    """Read the FieldList+IRR sheet -> set of normalized field names marked Y.
    Falls back to the built-in spec when the sheet is absent."""
    if "FieldList+IRR" in wb.sheetnames:
        ws = wb["FieldList+IRR"]
        y = set()
        for row in ws.iter_rows(values_only=True):
            if row and row[0] is not None and str(row[1]).strip().upper() == "Y":
                y.add(_norm(row[0]))
        if y:
            return y
    return set(NONMETRIC_Y) | {f"metric{i}-*" for i in range(1, 11)}

def _is_included(field_norm, irr_y):
    if field_norm in EXCLUDE_FREETEXT:
        return False
    if re.match(r"metric\d+-", field_norm):
        # any metric subfield is included if metrics are in the spec
        return any(k.startswith("metric") for k in irr_y)
    return field_norm in irr_y

# ----------------------------------------------------- Krippendorff alpha ----
def krippendorff_alpha(records, coders=None, paper=None):
    """Nominal Krippendorff's alpha over (paper,rq,field) units.
    records: (paper, rq, field, coder, value). Returns (alpha, n_units, n_pairs)."""
    units = defaultdict(dict)
    for p, q, f, c, v in records:
        if coders is not None and c not in coders:
            continue
        if paper is not None and p != paper:
            continue
        if v is None:
            continue
        units[(p, q, f)][c] = v

    coinc = defaultdict(float)
    n_units = 0
    for cv in units.values():
        vals = list(cv.values())
        m = len(vals)
        if m < 2:
            continue
        n_units += 1
        cnt = Counter(vals)
        for a in cnt:
            for b in cnt:
                pairs = cnt[a] * (cnt[a] - 1) if a == b else cnt[a] * cnt[b]
                coinc[(a, b)] += pairs / (m - 1)

    n_c = defaultdict(float)
    for (a, _b), o in coinc.items():
        n_c[a] += o
    n = sum(n_c.values())
    if n < 2:
        return None, n_units, n
    vals = list(n_c)
    Do = sum(coinc.get((a, b), 0.0) for a in vals for b in vals if a != b)
    De = sum(n_c[a] * n_c[b] for a in vals for b in vals if a != b) / (n - 1)
    if De == 0:
        return 1.0, n_units, n
    return 1 - Do / De, n_units, n

# ----------------------------------------------------------------- report ----
def coders_in(records):
    return sorted({c for *_x, c, _v in [(p, q, f, c, v) for p, q, f, c, v in records]})

def run(path, sheet, per_paper=False, pairs=False, per_field=False, csv_out=None,
        name_map=NAME_MAP):
    records, fields = load_workbook_records(path, sheet, name_map=name_map)
    allcoders = sorted({c for _p, _q, _f, c, _v in records})
    out_rows = []

    overall, nu, npr = krippendorff_alpha(records)
    print(f"\n=== IRR for sheet '{sheet}' ===")
    print(f"coders ({len(allcoders)}): {', '.join(allcoders)}")
    print(f"IRR fields used: {len(fields)} (free-text excluded, metrics normalized)")
    print(f"\nOVERALL alpha = {fmt(overall)}   (units={nu})")
    out_rows.append(["overall", "", fmt(overall), nu])

    if per_paper:
        print("\n-- per paper --")
        papers = sorted({p for p, *_ in records}, key=_papersort)
        for p in papers:
            a, u, _ = krippendorff_alpha(records, paper=p)
            print(f"   paper {p:>8}: alpha = {fmt(a):>7}  (units={u})")
            out_rows.append(["paper", p, fmt(a), u])

    if pairs:
        print("\n-- by coder pair --")
        for c1, c2 in combinations(allcoders, 2):
            a, u, _ = krippendorff_alpha(records, coders={c1, c2})
            if u == 0:
                continue
            print(f"   {c1:>10} & {c2:<10}: alpha = {fmt(a):>7}  (units={u})")
            out_rows.append([f"pair", f"{c1} & {c2}", fmt(a), u])

    if per_field:
        print("\n-- per field --")
        for f in fields:
            sub = [(p, q, ff, c, v) for p, q, ff, c, v in records if ff == f]
            a, u, _ = krippendorff_alpha(sub)
            print(f"   {f[:45]:45}: alpha = {fmt(a):>7}  (units={u})")
            out_rows.append(["field", f, fmt(a), u])

    if csv_out:
        with open(csv_out, "w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["scope", "key", "alpha", "n_units"])
            w.writerows(out_rows)
        print(f"\nwrote {csv_out}")

def _papersort(p):
    try:
        return (0, float(p))
    except (TypeError, ValueError):
        return (1, str(p))

def fmt(a):
    return "n/a" if a is None else f"{a:.4f}"

def main():
    ap = argparse.ArgumentParser(description="Compute Krippendorff IRR for a coding sheet.")
    ap.add_argument("workbook")
    ap.add_argument("--sheet", help="sheet/tab name (e.g. 52226)")
    ap.add_argument("--all-sheets", action="store_true",
                    help="run every sheet that has a Reviewer column")
    ap.add_argument("--per-paper", action="store_true")
    ap.add_argument("--pairs", action="store_true")
    ap.add_argument("--per-field", action="store_true")
    ap.add_argument("--csv", dest="csv_out")
    ap.add_argument("--aliases", metavar="PATH",
                    help='JSON {"spelling": "Canonical"} map of coder-name '
                         "aliases, merged over the built-in defaults")
    args = ap.parse_args()

    name_map = load_aliases(args.aliases) if args.aliases else NAME_MAP

    if args.all_sheets:
        wb = openpyxl.load_workbook(args.workbook, read_only=True)
        names = wb.sheetnames
        wb.close()
        for s in names:
            try:
                recs, _ = load_workbook_records(args.workbook, s, name_map=name_map)
            except SystemExit:
                continue
            if recs:
                run(args.workbook, s, args.per_paper, args.pairs, args.per_field,
                    name_map=name_map)
    else:
        if not args.sheet:
            ap.error("provide --sheet or --all-sheets")
        run(args.workbook, args.sheet, args.per_paper, args.pairs,
            args.per_field, args.csv_out, name_map=name_map)

if __name__ == "__main__":
    main()
